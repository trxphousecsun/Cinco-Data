import pandas as pd
import plotly.express as px
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import glob
import datetime

def process_excel_files(input_files, output_file, sheet_to_copy, company_name_column='Company Name'):
    # Create or load the output workbook
    if os.path.exists(output_file):
        output_wb = load_workbook(filename=output_file)
        output_ws = output_wb.create_sheet(sheet_to_copy) if sheet_to_copy not in output_wb.sheetnames else output_wb[sheet_to_copy]
    else:
        output_wb = Workbook()
        output_ws = output_wb.create_sheet(sheet_to_copy)

    current_row = 1
    image_row = 1

    for input_file in input_files:
        try:
            # Load the input workbook
            input_wb = load_workbook(filename=input_file)
        except Exception as e:
            print(f"Error opening file {input_file}: {e}")
            continue

        if sheet_to_copy not in input_wb.sheetnames:
            print(f"Sheet '{sheet_to_copy}' not found in {input_file}.")
            continue

        df = pd.read_excel(input_file, sheet_name=sheet_to_copy)
        print(f"Columns in {input_file}: {df.columns.tolist()}")  # Debugging statement

        if company_name_column not in df.columns:
            print(f"Column '{company_name_column}' not found in {input_file}.")
            continue
        company_name = df[company_name_column].iloc[0]

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=(current_row == 1))):
            for c_idx, value in enumerate(row, 1):
                output_ws.cell(row=current_row+r_idx, column=c_idx, value=value)
        current_row += len(df) + 2

        for column, metric in [('NWC (Thousands/Millions)', 'NWC'), ('FCFF (Thousands/Millions)', 'FCFF')]:
            fig = px.line(df, x='Filing Year', y=column, markers=True, title=f'{company_name} {metric} Time Series')
            fig.update_layout(xaxis_title='Year', yaxis_title=column, plot_bgcolor='white')
            fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='lightgrey')
            fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='lightgrey')
            img_path = f'/tmp/{company_name}_{metric}_{datetime.datetime.now().strftime("%Y%m%d%H%M%S")}.png'
            fig.write_image(img_path)
            
            img = Image(img_path)
            output_ws.add_image(img, f'H{image_row}')
            image_row += 20

    output_wb.save(filename=output_file)

def get_excel_files_in_directory(directory_path):
    return glob.glob(os.path.join(directory_path, '*.xlsx'))

# Example usage
directory_path = '/Users/jazzhashzzz/Desktop/testfiles/tech'
input_excel_files = get_excel_files_in_directory(directory_path)
output_excel_file = '/Users/jazzhashzzz/Desktop/testfiles/tech/teckkyy.xlsx'
sheet_name = 'Calculated Data'

process_excel_files(input_excel_files, output_excel_file, sheet_name)
